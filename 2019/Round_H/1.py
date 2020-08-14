import os
import openpyxl
import pandas as pd
import re

path = 'C:/Users/User/Desktop/工作文件/1.实施中项目/1.阿里马来MY88-M109/MY88-M109/4autocollect-device-status/20200730'

DSW = [
            {'Brand'   :'H3C',
             'BGP'     :'Peers in established state: 32',
             'LLDP'    :32,
             'Version' :['S12500X_T-CMW710-SYSTEM-R3203P01.bin',
                         'S12500X_T-CMW710-SYSTEM-R3203P01H03.bin',
                         'S12500X_T-CMW710-SYSTEM-WEAK-R3203P01H02.bin'],
             'Power'   :'Faulty',
             'Fan'     :'Absent'
            }, 

            {'Brand'   :'Ruijie',
             'BGP'     :'established neighbors 32',
             'LLDP'    :32,
             'Version' :['N18000-X_RGOS 11.3(1)B6P38',
                         'SP2(running)'],
             'Power'   :'fail',
             'Fan'     :''
            },

            {'Brand'   :'Cisco',
             'BGP'     :'capable peers 32',
             'LLDP'    :32,
             'Version' :['9.2(3y)'],
             'Power'   :'0 W',
             'Fan'     :'Absent'
            },

            {'Brand'   :'Huawei',
             'BGP'     :'',
             'LLDP'    :'',
             'Version' :[''],
             'Power'   :'',
             'Fan'     :''
            }
          ]

PSW = [
            {'Brand'   :'H3C',
             'BGP'     :'',
             'LLDP'    :'',
             'Version' :['s9820-cmw710-system-r6555p01.bin',
                         'S9820-CMW710-SYSTEM-R6555P01H09.bin',
                         'S9820-CMW710-SYSTEM-WEAK-R6555P01H02.bin'],
             'Power'   :'Fault',
             'Fan'     :'Absent'
            }, 

            {'Brand'   :'Ruijie',
             'BGP'     :'',
             'LLDP'    :'',
             'Version' :['S6500_RGOS 11.0(5)B9P43S',
                         'SP1(running)'],
             'Power'   :'disconnect',
             'Fan'     :'disconnect'
            },

            {'Brand'   :'Cisco',
             'BGP'     :'',
             'LLDP'    :'',
             'Version' :['9.2(3y)'],
             'Power'   :'0 W',
             'Fan'     :'Absent'
            },

            {'Brand'   :'Huawei',
             'BGP'     :'',
             'LLDP'    :'',
             'Version' :['CE8850EI V200R005C10SPC800',
                         'V200R005SPH013'],
             'Power'   :'NO',
             'Fan'     :'NO'
            }
          ]

ASW = [
            {'Brand'   :'H3C',
             'BGP'     :'Peers in established state: 8',
             'LLDP'    :8,
             'Version' :['s9850_6850-cmw710-system-r6555p01.bin',
                         'S9850_6850-CMW710-SYSTEM-R6555P01H21.bin',
                         'S9850_6850-CMW710-SYSTEM-WEAK-R6555P01H02.bin',
                         'S9850_6850-CMW710-SYSTEM-WEAK-R6555P01H04.bin'],
             'Power'   :'Fault',
             'Fan'     :'Absent'
            }, 

            {'Brand'   :'Ruijie',
             'BGP'     :'established neighbors 8',
             'LLDP'    :8,
             'Version' :['S6500_RGOS 11.0(5)B9P43',
                         'SP2(running)'],
             'Power'   :'disconnect',
             'Fan'     :'disconnect'
            },

            {'Brand'   :'Cisco',
             'BGP'     :'capable peers 8',
             'LLDP'    :8,
             'Version' :['9.2(3y)'],
             'Power'   :'0 W',
             'Fan'     :'Absent'
            },

            {'Brand'   :'Huawei',
             'BGP'     :'Peers in established state : 8',
             'LLDP'    :8,
             'Version' :['CE6865EI V200R005C10SPC800',
                         'V200R005SPH013'],
             'Power'   :'NO',
             'Fan'     :'NO'
            }
          ]

wb = openpyxl.Workbook()
s1 = wb.active
s2 = wb.create_sheet()
s3 = wb.create_sheet()
s4 = wb.create_sheet()
s5 = wb.create_sheet()

s1.title = 'BGP测试报告'
s1['A1'] = '机器名字'
s1['B1'] = '机器错误报告'

s2.title = 'LLDP测试报告'
s2['A1'] = '机器名字'
s2['B1'] = '机器错误报告'

s3.title = 'Version测试报告'
s3['A1'] = '机器名字'
s3['B1'] = '机器错误报告'

s4.title = 'Power测试报告'
s4['A1'] = '机器名字'
s4['B1'] = '机器错误报告'

s5.title = 'Fan测试报告'
s5['A1'] = '机器名字'
s5['B1'] = '机器错误报告'
        
for subdir in os.listdir(path):
    if subdir.startswith('ASW'):
        for file in os.listdir(os.path.join(path,subdir)):
            device_name = subdir
            for read in ASW:
                with open(os.path.join(path,subdir,'device_version.txt')) as a:
                    if read.get('Brand') in a.read():
                        if file.endswith('BGP.txt'):
                        with open(os.path.join(path,subdir,'device_BGP.txt')) as b:
                            if read.get('BGP') in b.read():
                                continue
                            else:
                                s1.append([str(device_name),'BGP Error'])
                                continue
                        elif file.endswith('LLDP.txt'):
                            with open(os.path.join(path,subdir,'device_LLDP.txt')) as b:
                                ln = b.read()
                                if ln.count('PSW') == read.get('LLDP'):
                                    continue
                                else:
                                    s2.append([str(device_name),'LLDP Error'])
                                    continue
                        elif file.endswith('version.txt'):
                            with open(os.path.join(path,subdir,'device_version.txt')) as b:
                                ln = b.read()
                                for c in read.get('Version'):
                                    if c in ln:
                                        continue
                                    else:
                                        s3.append([str(device_name),'Version Error'])
                        elif file.endswith('enviroment.txt'):
                            with open(os.path.join(path,subdir,'show_enviroment.txt')) as b:
                                if read.get('Power') in b.read():
                                    s4.append([str(device_name),'Power Error'])
                                else:
                                    continue
                        elif file.endswith('Fan.txt'):
                            with open(os.path.join(path,subdir,'device_Fan.txt')) as b:
                                if read.get('Fan') in b.read():
                                    s5.append([str(device_name),'Fan Error'])
                                else:
                                    continue
    elif subdir.startswith('PSW'):
        for file in os.listdir(os.path.join(path,subdir)):
            device_name = subdir
            for read in PSW:
                with open(os.path.join(path,subdir,'device_version.txt')) as a:
                    if read.get('Brand') in a.read():
                        if file.endswith('version.txt'):
                            with open(os.path.join(path,subdir,'device_version.txt')) as b:
                                ln = b.read()
                                for c in read.get('Version'):
                                    if c in ln:
                                        continue
                                    else:
                                        s3.append([str(device_name),'Version Error'])
                        elif file.endswith('enviroment.txt'):
                            with open(os.path.join(path,subdir,'show_enviroment.txt')) as b:
                                if read.get('Power') in b.read():
                                    s4.append([str(device_name),'Power Error'])
                                else:
                                    continue
                        elif file.endswith('Fan.txt'):
                            with open(os.path.join(path,subdir,'device_Fan.txt')) as b:
                                if read.get('Fan') in b.read():
                                    s5.append([str(device_name),'Fan Error'])
                                else:
                                    continue
    elif subdir.startswith('DSW'):
        for file in os.listdir(os.path.join(path,subdir)):
            device_name = subdir
            for read in DSW:
                with open(os.path.join(path,subdir,'device_version.txt')) as a:
                    if read.get('Brand') in a.read():
                        if file.endswith('BGP.txt'):
                            with open(os.path.join(path,subdir,'device_BGP.txt')) as b:
                                if read.get('BGP') in b.read():
                                    continue
                                else:
                                    s1.append([str(device_name),'BGP Error'])
                                    continue
                        elif file.endswith('LLDP.txt'):
                            with open(os.path.join(path,subdir,'device_LLDP.txt')) as b:
                                ln = b.read()
                                if ln.count('PSW') == read.get('LLDP'):
                                    continue
                                else:
                                    s2.append([str(device_name),'LLDP Error'])
                                    continue
                        elif file.endswith('version.txt'):
                            with open(os.path.join(path,subdir,'device_version.txt')) as b:
                                ln = b.read()
                                for c in read.get('Version'):
                                    if c in ln:
                                        continue
                                    else:
                                        s3.append([str(device_name),'Version Error'])
                        elif file.endswith('enviroment.txt'):
                            with open(os.path.join(path,subdir,'show_enviroment.txt')) as b:
                                if read.get('Power') in b.read():
                                    s4.append([str(device_name),'Power Error'])
                                else:
                                    continue
                        elif file.endswith('Fan.txt'):
                            with open(os.path.join(path,subdir,'device_Fan.txt')) as b:
                                if read.get('Fan') in b.read():
                                    s5.append([str(device_name),'Fan Error'])
                                else:
                                    continue
wb.save('互联检查结果.xlsx')
wb.close()